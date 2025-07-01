from flask import Flask, request, jsonify
from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta
import os
from dateutil.relativedelta import relativedelta

app = Flask(__name__)

# Database file
DB_FILE = 'audit_allocations.xlsx'

# Initialize database if not exists
if not os.path.exists(DB_FILE):
    wb = Workbook()
    # Members sheet
    ws_members = wb.active
    ws_members.title = "Members"
    ws_members.append(["MemberID", "Name", "Email", "Department"])

    # Add some sample members  
    sample_members = [  
        [1, "John Doe", "john@example.com", "Audit"],  
        [2, "Jane Smith", "jane@example.com", "Tax"],  
        [3, "Mike Johnson", "mike@example.com", "Advisory"],  
        [4, "Sarah Williams", "sarah@example.com", "Audit"],  
        [5, "David Brown", "david@example.com", "Tax"]  
    ]  
    for member in sample_members:  
        ws_members.append(member)  
      
    # Allocations sheet  
    ws_alloc = wb.create_sheet("Allocations")  
    ws_alloc.append(["AllocationID", "MemberID", "Description", "Section", "FromDate", "ToDate", "FinancialYear"])  
      
    # Add some sample allocations  
    today = datetime.now().date()  
    
    def get_financial_year_sample(date_str):
        """Get financial year (April-March) for a given date"""
        date = datetime.strptime(date_str, '%Y-%m-%d').date()
        if date.month >= 4:
            return f"{date.year}-{date.year+1}"
        return f"{date.year-1}-{date.year}"
    
    sample_allocations = [  
        [1, 1, "Annual Financial Audit", "Section 1", (today - timedelta(days=30)).isoformat(), (today - timedelta(days=15)).isoformat(), get_financial_year_sample((today - timedelta(days=30)).isoformat())],  
        [2, 2, "Tax Compliance Review", "Section 2", (today - timedelta(days=10)).isoformat(), (today + timedelta(days=5)).isoformat(), get_financial_year_sample((today - timedelta(days=10)).isoformat())],  
        [3, 3, "Internal Controls Assessment", "Section 3", (today + timedelta(days=5)).isoformat(), (today + timedelta(days=15)).isoformat(), get_financial_year_sample((today + timedelta(days=5)).isoformat())],  
        [4, 4, "SOX Compliance", "Section 4", (today + timedelta(days=20)).isoformat(), (today + timedelta(days=30)).isoformat(), get_financial_year_sample((today + timedelta(days=20)).isoformat())]  
    ]  
    for alloc in sample_allocations:  
        ws_alloc.append(alloc)  
      
    wb.save(DB_FILE)

def get_financial_year(date_str):
    """Get financial year (April-March) for a given date"""
    date = datetime.strptime(date_str, '%Y-%m-%d').date()
    if date.month >= 4:
        return f"{date.year}-{date.year+1}"
    return f"{date.year-1}-{date.year}"

def is_available(member_id, from_date, to_date, exclude_allocation_id=None):
    """Check if member is available for given date range"""
    wb = load_workbook(DB_FILE)
    ws = wb["Allocations"]

    for row in ws.iter_rows(min_row=2, values_only=True):  
        if row[1] == member_id:  
            # Skip if this is the allocation being excluded (for updates)
            if exclude_allocation_id and row[0] == exclude_allocation_id:
                continue
                
            alloc_from = datetime.strptime(row[4], '%Y-%m-%d').date() if isinstance(row[4], str) else row[4].date()  
            alloc_to = datetime.strptime(row[5], '%Y-%m-%d').date() if isinstance(row[5], str) else row[5].date()  
              
            # Check for date overlap  
            if not (to_date < alloc_from or from_date > alloc_to):  
                return False  
    return True

@app.route('/api/members', methods=['GET'])
def get_members():
    """Get all members"""
    wb = load_workbook(DB_FILE)
    ws = wb["Members"]
    members = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        members.append({
            "id": row[0],
            "name": row[1],
            "email": row[2],
            "department": row[3]
        })
    return jsonify(members)

@app.route('/api/available-members', methods=['POST'])
def get_available_members():
    """Get members available for given date range"""
    data = request.json
    from_date = datetime.strptime(data['fromDate'], '%Y-%m-%d').date()
    to_date = datetime.strptime(data['toDate'], '%Y-%m-%d').date()

    wb = load_workbook(DB_FILE)  
    ws_members = wb["Members"]  
    available_members = []  
      
    for row in ws_members.iter_rows(min_row=2, values_only=True):  
        if is_available(row[0], from_date, to_date):  
            available_members.append({  
                "id": row[0],  
                "name": row[1]  
            })  
      
    return jsonify(available_members)

@app.route('/api/suggest-dates', methods=['POST'])
def suggest_dates():
    """Suggest available date ranges for a member"""
    data = request.json
    member_id = data['memberId']
    days_required = int(data['daysRequired'])

    wb = load_workbook(DB_FILE)  
    ws = wb["Allocations"]  
      
    # Get all allocations for this member, sorted by date  
    member_allocations = []  
    for row in ws.iter_rows(min_row=2, values_only=True):  
        if row[1] == member_id:  
            alloc_from = datetime.strptime(row[4], '%Y-%m-%d').date() if isinstance(row[4], str) else row[4].date()  
            alloc_to = datetime.strptime(row[5], '%Y-%m-%d').date() if isinstance(row[5], str) else row[5].date()  
            member_allocations.append((alloc_from, alloc_to))  
      
    # Sort allocations by start date  
    member_allocations.sort()  
      
    today = datetime.now().date()  
    suggestions = []  
      
    # Check availability before first allocation  
    if len(member_allocations) > 0:  
        first_alloc_start = member_allocations[0][0]  
        available_days = (first_alloc_start - today).days  
        if available_days >= days_required:  
            suggestions.append({  
                "startDate": today.isoformat(),  
                "endDate": (today + timedelta(days=days_required - 1)).isoformat(),  
                "availableDays": available_days  
            })  
      
    # Check gaps between allocations  
    for i in range(len(member_allocations) - 1):  
        current_end = member_allocations[i][1]  
        next_start = member_allocations[i+1][0]  
          
        if next_start > current_end:  
            gap_days = (next_start - current_end).days - 1  
            if gap_days >= days_required:  
                start_date = current_end + timedelta(days=1)  
                suggestions.append({  
                    "startDate": start_date.isoformat(),  
                    "endDate": (start_date + timedelta(days=days_required - 1)).isoformat(),  
                    "availableDays": gap_days  
                })  
      
    # Check availability after last allocation  
    if len(member_allocations) > 0:  
        last_alloc_end = member_allocations[-1][1]  
        start_date = last_alloc_end + timedelta(days=1)  
        suggestions.append({  
            "startDate": start_date.isoformat(),  
            "endDate": (start_date + timedelta(days=days_required - 1)).isoformat(),  
            "availableDays": 365  # Arbitrary large number  
        })  
    else:  
        # No allocations at all - suggest starting today  
        suggestions.append({  
            "startDate": today.isoformat(),  
            "endDate": (today + timedelta(days=days_required - 1)).isoformat(),  
            "availableDays": 365  
        })  
      
    # Sort suggestions by start date  
    suggestions.sort(key=lambda x: x['startDate'])  
      
    return jsonify(suggestions[:5])  # Return top 5 suggestions

@app.route('/api/allocations', methods=['GET'])
def get_allocations():
    """Get all allocations for a financial year"""
    financial_year = request.args.get('financialYear')

    wb = load_workbook(DB_FILE)  
    ws = wb["Allocations"]  
    allocations = []  
      
    for row in ws.iter_rows(min_row=2, values_only=True):  
        row_fy = row[6] if isinstance(row[6], str) else get_financial_year(row[4].strftime('%Y-%m-%d'))  
        if row_fy == financial_year:  
            allocations.append({  
                "id": row[0],  
                "memberId": row[1],  
                "description": row[2],  
                "section": row[3],  
                "fromDate": row[4].strftime('%Y-%m-%d') if not isinstance(row[4], str) else row[4],  
                "toDate": row[5].strftime('%Y-%m-%d') if not isinstance(row[5], str) else row[5],  
                "financialYear": row_fy  
            })  
      
    return jsonify(allocations)

@app.route('/api/allocations', methods=['POST'])
def create_allocation():
    """Create new allocation"""
    data = request.json
    today = datetime.now().date()
    from_date = datetime.strptime(data['fromDate'], '%Y-%m-%d').date()
    to_date = datetime.strptime(data['toDate'], '%Y-%m-%d').date()

    # Validate dates  
    if from_date > to_date:  
        return jsonify({"error": "From date cannot be after to date"}), 400  
      
    # Check if member is available  
    if not is_available(data['memberId'], from_date, to_date):  
        return jsonify({"error": "Member is not available for the selected dates"}), 400  
      
    wb = load_workbook(DB_FILE)  
    ws = wb["Allocations"]  
      
    # Generate new ID  
    existing_ids = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]  
    new_id = max(existing_ids) + 1 if existing_ids else 1  
      
    # Add new allocation  
    financial_year = get_financial_year(data['fromDate'])  
    ws.append([  
        new_id,  
        data['memberId'],  
        data['description'],  
        data['section'],  
        data['fromDate'],  
        data['toDate'],  
        financial_year  
    ])  
      
    wb.save(DB_FILE)  
    return jsonify({"success": True, "id": new_id})

@app.route('/api/allocations/<int:alloc_id>', methods=['PUT'])
def update_allocation(alloc_id):
    """Update existing allocation"""
    data = request.json
    today = datetime.now().date()
    from_date = datetime.strptime(data['fromDate'], '%Y-%m-%d').date()
    to_date = datetime.strptime(data['toDate'], '%Y-%m-%d').date()

    # Validate dates  
    if from_date > to_date:  
        return jsonify({"error": "From date cannot be after to date"}), 400  
      
    wb = load_workbook(DB_FILE)  
    ws = wb["Allocations"]  
      
    # Find the allocation  
    for row in ws.iter_rows(min_row=2):  
        if row[0].value == alloc_id:  
            original_from = datetime.strptime(row[4].value, '%Y-%m-%d').date() if isinstance(row[4].value, str) else row[4].value.date()  
            original_to = datetime.strptime(row[5].value, '%Y-%m-%d').date() if isinstance(row[5].value, str) else row[5].value.date()  
              
            # Validate: Cannot move allocation to the past  
            if from_date < today and from_date != original_from:  
                return jsonify({"error": "Cannot move allocation to the past"}), 400  
              
            # Validate: Cannot shorten allocation to before today for current allocations  
            if original_from <= today <= original_to:  
                if from_date > original_from:  
                    return jsonify({"error": "Cannot move start date forward for current allocation"}), 400  
                if to_date < today:  
                    return jsonify({"error": "Cannot set end date before today for current allocation"}), 400  
              
            # Check if member is available for new dates (excluding this allocation)  
            if not is_available(row[1].value, from_date, to_date, exclude_allocation_id=alloc_id):  
                return jsonify({"error": "Member is not available for the selected dates"}), 400  
              
            # Update the allocation  
            row[2].value = data['description']  # Description  
            row[3].value = data['section']      # Section  
            row[4].value = data['fromDate']     # From Date  
            row[5].value = data['toDate']       # To Date  
            break  
      
    wb.save(DB_FILE)  
    return jsonify({"success": True})

@app.route('/api/allocations/<int:alloc_id>', methods=['DELETE'])
def delete_allocation(alloc_id):
    """Delete an allocation"""
    today = datetime.now().date()

    wb = load_workbook(DB_FILE)  
    ws = wb["Allocations"]  
      
    # Find the allocation  
    for row in ws.iter_rows(min_row=2):  
        if row[0].value == alloc_id:  
            from_date = datetime.strptime(row[4].value, '%Y-%m-%d').date() if isinstance(row[4].value, str) else row[4].value.date()  
              
            # Validate: Cannot delete past allocations  
            if from_date < today:  
                return jsonify({"error": "Cannot delete past allocations"}), 400  
              
            # Delete the row  
            ws.delete_rows(row[0].row)  
            break  
      
    wb.save(DB_FILE)  
    return jsonify({"success": True})


 



from flask import send_file
from fpdf import FPDF
import io

MANAGER_FILE = 'managers.txt'  # Each line = one manager

class PDF(FPDF):
    def header(self):
        self.image('static/logo.png', 10, 8, 20)
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, 'INTER OFFICE MEMORANDUM', border=0, ln=1, align='C')
        self.set_font('Arial', '', 10)
        self.cell(0, 8, "From: IAD", ln=1)
        self.cell(0, 8, "To: CS", ln=1)
        self.cell(0, 8, "Ref: __/2025-26/IAD", ln=1)
        self.cell(0, 8, "Date: 27.06.2025", ln=1)
        self.ln(5)

    def table_header(self):
        headers = ['Sl No', 'Audit Ref No', 'UNIT/MODULE', 'AUDITORS', 'AUDIT MGR', 'From', 'To', 'Remarks']
        widths = [10, 20, 50, 30, 25, 20, 20, 35]
        self.set_font('Arial', 'B', 9)
        for header, width in zip(headers, widths):
            self.cell(width, 8, header, border=1)
        self.ln()

    def table_row(self, index, row):
        values = [
            str(index),
            str(row[0]),  # AllocationID
            row[2],       # Description
            row[-2],      # Filtered Auditors
            row[3],       # Audit Manager
            row[4],       # FromDate
            row[5],       # ToDate
            row[6] if len(row) > 6 else ""  # Remarks
        ]
        widths = [10, 20, 50, 30, 25, 20, 20, 35]
        self.set_font('Arial', '', 8)
        for val, width in zip(values, widths):
            self.cell(width, 8, val, border=1)
        self.ln()

def classify_unit(description):
    description = description.lower()
    if description.startswith('ho-'):
        return 'HO Units'
    elif description.startswith('sof-'):
        return 'SOF Units'
    return 'Plant Units'

@app.route('/generate-pdf', methods=['GET'])
def generate_pdf():
    financial_year = request.args.get('financialYear')
    if not financial_year:
        return jsonify({"error": "Missing financial year"}), 400

    wb = load_workbook(DB_FILE)
    ws = wb["Allocations"]

    if not os.path.exists(MANAGER_FILE):
        return jsonify({"error": "Manager file not found"}), 404

    with open(MANAGER_FILE, 'r') as f:
        managers = [line.strip() for line in f if line.strip()]

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_fy = row[6] if isinstance(row[6], str) else get_financial_year(row[4])
        if row_fy != financial_year:
            continue

        description = row[2]
        unit_type = classify_unit(description)
        audit_mgr = row[3]
        filtered_auditors = ', '.join([
            name.strip() for name in row[2].split(',') if name.strip() not in managers
        ])
        remarks = row[6] if len(row) > 6 else ""
        data.append({
            "unit": unit_type,
            "manager": audit_mgr,
            "row": list(row) + [filtered_auditors, remarks]
        })

    if not data:
        return jsonify({"error": "No data found for selected financial year"}), 404

    pdf = PDF()
    pdf.add_page()

    for unit_type in ["HO Units", "SOF Units", "Plant Units"]:
        pdf.set_font('Arial', 'B', 11)
        pdf.cell(0, 10, unit_type, ln=1)
        unit_data = [d for d in data if d["unit"] == unit_type]
        managers_seen = set()

        index = 1
        for entry in unit_data:
            manager = entry["manager"]
            if manager not in managers_seen:
                pdf.set_font('Arial', 'B', 10)
                pdf.cell(0, 8, f"Manager: {manager}", ln=1)
                pdf.table_header()
                managers_seen.add(manager)
                index = 1

            pdf.table_row(index, entry["row"])
            index += 1

    buffer = io.BytesIO()
    pdf.output(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"audit_programme_{financial_year}.pdf",
        mimetype='application/pdf'
    )

member_ids = [int(mid.strip()) for mid in str(member_ids_raw).split(",") if mid.strip().isdigit()]
        member_names = [member_map.get(mid, f"ID:{mid}") for mid in member_ids]

        audit_mgr = next((name for name in member_names if name in managers), "")
        auditors = " / ".join([name for name in member_names if name != audit_mgr])

        unit_type = classify_unit(description)

        data.append({
            "unit": unit_type,
            "row": [allocation_id, allocation_id, description, auditors, audit_mgr, from_date, to_date, ""]
        })

    if not data:
        return jsonify({"error": "No data found for selected financial year"}), 404

    pdf = PDF(header_date=header_date, ref_year=financial_year)
    pdf.add_page()

    for unit_type in ["HO Units", "SOF Units", "Plant Units"]:
        unit_data = [d for d in data if d["unit"] == unit_type]
        if not unit_data:
            continue

        pdf.set_font('Arial', 'B', 11)
        pdf.cell(0, 10, unit_type, ln=1)
        pdf.table_header()

        for idx, entry in enumerate(unit_data, start=1):
            pdf.table_row(idx, entry["row"])

        pdf.ln(5)

    buffer = io.BytesIO()
    pdf_output = pdf.output(dest='S').encode('latin1')
    buffer.write(pdf_output)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"audit_programme_{financial_year}.pdf",
        mimetype='application/pdf'
    )

if __name__ == '__main__':
    app.run(debug=True)


from flask import Flask, request, jsonify, send_file
from openpyxl import load_workbook
from fpdf import FPDF
import io
import os
from flask_cors import CORS
from datetime import datetime
from collections import defaultdict

app = Flask(__name__)
CORS(app)
MANAGER_FILE = 'managers.txt'
DB_FILE = 'audit_allocations_1.xlsx'

class PDF(FPDF):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.quarter_info = None
        self.fin_year = None
        self.header_date = None

    def header(self):
        # Logo
        self.image('static/logo.png', 10, 8, 20)
        
        # Company name under logo (left side)
        self.set_font('Arial', 'B', 12)
        self.set_xy(10, 30)
        self.cell(30, 8, 'XYZ')
        
        # Main heading parallel to company name (center-right)
        self.set_font('Arial', 'B', 14)
        self.set_xy(70, 30)
        self.cell(0, 8, 'INTER OFFICE MEMORANDUM', align='C')
        
        # Horizontal line after logo and headings
        self.set_xy(10, 45)
        self.line(10, 45, 200, 45)
        
        # From and Ref parallel (same line)
        self.set_font('Arial', '', 10)
        self.set_xy(10, 55)
        self.cell(80, 8, 'From: IAD')
        
        self.set_xy(120, 55)
        self.cell(0, 8, f'Ref: __/{self.fin_year}/IAD', align='R')
        
        # To and Date parallel (same line)
        self.set_xy(10, 70)
        self.cell(80, 8, 'To: CS')
        
        self.set_xy(120, 70)
        self.cell(0, 8, f'Date: {self.header_date}', align='R')
        
        # Add line under To and Date
        self.line(10, 78, 200, 78)
        
        # Subject line with dynamic quarter
        self.set_xy(10, 85)
        self.set_font('Arial', 'B', 12)
        self.cell(0, 8, f'AUDIT PROGRAMME FOR THE {self.quarter_info}', ln=1, align='C')
        
        self.ln(10)

    def set_header_data(self, fin_year, header_date, quarter_info):
        self.fin_year = fin_year
        self.header_date = header_date
        self.quarter_info = quarter_info

    def table_header(self):
        # Main table headers with proper widths
        self.set_font('Arial', 'B', 9)
        
        # Column widths - adjusted to fit page width (210mm - 20mm margins = 190mm)
        widths = [12, 18, 50, 35, 20, 20, 20, 25]  # Total: 200mm
        
        # Draw first row of headers
        start_x = 10
        start_y = self.get_y()
        
        # Draw individual cells for first row
        headers_row1 = ['Sl', 'Audit Ref', 'UNIT/MODULE', 'AUDITORS', 'AUDIT', 'Audit Period', 'REMARKS']
        header_widths = [12, 18, 50, 35, 20, 40, 25]  # Audit Period spans 2 columns (20+20=40)
        
        for i, (header, width) in enumerate(zip(headers_row1, header_widths)):
            self.set_xy(start_x + sum(header_widths[:i]), start_y)
            if i < 5:  # Regular columns
                self.cell(width, 16, header, border=1, align='C')
            elif i == 5:  # Audit Period column - only top, left, right borders
                self.cell(width, 8, header, border=1, align='C')
            else:  # REMARKS column
                self.cell(width, 16, header, border=1, align='C')
        
        # Draw second row for sub-headers
        sub_headers = ['No', 'No.', '', '', 'MGR', 'From', 'To', '']
        
        for i, (sub_header, width) in enumerate(zip(sub_headers, widths)):
            self.set_xy(start_x + sum(widths[:i]), start_y + 8)
            if i < 5:  # First 5 columns get 8px height for second row
                self.cell(width, 8, sub_header, border=1, align='C')
            elif i in [5, 6]:  # From and To columns
                self.cell(width, 8, sub_header, border=1, align='C')
            else:  # REMARKS column - empty cell for second row
                self.cell(width, 8, sub_header, border=1, align='C')
        
        self.set_xy(start_x, start_y + 16)

    def add_section_header(self, section_name):
        """Add section headers like 'HO Units', 'Plant Units' etc."""
        self.set_font('Arial', 'B', 10)
        self.ln(5)
        self.cell(0, 8, section_name, ln=1, align='C')
        self.ln(2)
        
    def get_text_height(self, text, width, font_size=8):
        """Calculate the height needed for text within given width"""
        self.set_font('Arial', '', font_size)
        char_width = font_size * 0.35  # More accurate character width
        chars_per_line = max(1, int((width - 4) / char_width))
        
        lines_needed = max(1, (len(str(text)) + chars_per_line - 1) // chars_per_line)
        return max(12, lines_needed * (font_size + 1))

    def multi_cell_row(self, data, widths, min_height=12):
        """Create a row with multi-line cells that align properly"""
        max_height = min_height
        font_size = 8
        
        # Calculate maximum height needed
        for text, width in zip(data, widths):
            if len(str(text)) > 0:
                text_height = self.get_text_height(str(text), width, font_size)
                max_height = max(max_height, text_height)
        
        start_x = 10
        start_y = self.get_y()
        
        # Draw each cell
        for i, (text, width) in enumerate(zip(data, widths)):
            self.set_xy(start_x + sum(widths[:i]), start_y)
            
            # Draw cell border
            self.rect(self.get_x(), self.get_y(), width, max_height)
            
            # Add text content
            self.set_font('Arial', '', font_size)
            if i in [0, 1, 4, 5, 6]:  # Center aligned columns
                self.set_xy(start_x + sum(widths[:i]), start_y + (max_height - font_size) / 2)
                self.cell(width, font_size, str(text), align='C')
            else:  # Left aligned columns with padding
                self.set_xy(start_x + sum(widths[:i]) + 2, start_y + 2)
                # For long text, use multi_cell
                if len(str(text)) > (width - 4) / (font_size * 0.35):
                    # Calculate available height for multi-cell
                    available_height = max_height - 4
                    line_height = font_size + 1
                    self.multi_cell(width - 4, line_height, str(text), align='L')
                else:
                    self.cell(width - 4, font_size, str(text), align='L')
        
        # Move to next row
        self.set_xy(start_x, start_y + max_height)

    def table_row(self, index, row, audit_ref_no=''):
        data = [
            str(index),
            str(audit_ref_no),
            str(row['Description']),
            str(row['Auditors']),
            str(row['Manager']),
            str(row['FromDate']),
            str(row['ToDate']),
            str(row.get('Remarks', ''))
        ]
        
        # Column widths matching header
        widths = [12, 18, 50, 35, 20, 20, 20, 25]
        
        # Check if we need a new page
        if self.get_y() > 250:
            self.add_page()
            self.table_header()
        
        self.multi_cell_row(data, widths)

def calculate_quarter_from_dates(grouped_data):
    """Calculate quarter and year based on the allocation dates"""
    if not grouped_data:
        return "QUARTER NOT DETERMINED"
    
    # Get all from_dates to determine the dominant quarter
    from_dates = []
    for (desc, from_date, to_date), val in grouped_data.items():
        if isinstance(from_date, datetime):
            from_dates.append(from_date)
        elif isinstance(from_date, str):
            try:
                # Try to parse string date
                parsed_date = datetime.strptime(from_date, '%Y-%m-%d')
                from_dates.append(parsed_date)
            except ValueError:
                try:
                    parsed_date = datetime.strptime(from_date, '%d-%m-%Y')
                    from_dates.append(parsed_date)
                except ValueError:
                    continue
    
    if not from_dates:
        return "QUARTER NOT DETERMINED"
    
    # Find the most common quarter
    quarter_counts = {1: 0, 2: 0, 3: 0, 4: 0}
    years = []
    
    for date in from_dates:
        years.append(date.year)
        month = date.month
        if month in [4, 5, 6]:
            quarter_counts[1] += 1
        elif month in [7, 8, 9]:
            quarter_counts[2] += 1
        elif month in [10, 11, 12]:
            quarter_counts[3] += 1
        else:  # [1, 2, 3]
            quarter_counts[4] += 1
    
    # Get the most common quarter
    dominant_quarter = max(quarter_counts, key=quarter_counts.get)
    
    # Get the most common year
    most_common_year = max(set(years), key=years.count) if years else datetime.now().year
    
    # Generate quarter text
    quarter_texts = {
        1: f"QUARTER APRIL - JUNE {most_common_year}",
        2: f"QUARTER JULY - SEPTEMBER {most_common_year}",
        3: f"QUARTER OCTOBER - DECEMBER {most_common_year}",
        4: f"QUARTER JANUARY - MARCH {most_common_year + 1}"
    }
    
    return quarter_texts.get(dominant_quarter, "QUARTER NOT DETERMINED")

@app.route('/generate-pdf', methods=['POST'])
def generate_pdf():
    try:
        fin_year = request.args.get('financialYear')
        header_date = request.json.get('headerDate')

        if not fin_year or not header_date:
            return jsonify({'error': 'Missing financialYear or headerDate'}), 400

        if not os.path.exists(DB_FILE) or not os.path.exists(MANAGER_FILE):
            return jsonify({'error': 'Required files not found'}), 404

        wb = load_workbook(DB_FILE)
        ws_alloc = wb['Allocations']
        ws_members = wb['Members']

        with open(MANAGER_FILE, 'r') as f:
            managers = [line.strip() for line in f if line.strip()]

        member_map = {row[0]: row[1] for row in ws_members.iter_rows(min_row=2, values_only=True)}

        # Group by (description, from_date, to_date)
        grouped = defaultdict(lambda: {'Auditors': [], 'Manager': None})

        for row in ws_alloc.iter_rows(min_row=2, values_only=True):
            alloc_id, member_id, desc, from_date, to_date, year = row
            if str(year) != fin_year:
                continue

            key = (desc, from_date, to_date)
            name = member_map.get(member_id, 'Unknown')

            if name in managers:
                grouped[key]['Manager'] = name
            else:
                grouped[key]['Auditors'].append(name)

        # Calculate quarter based on allocation data
        quarter_info = calculate_quarter_from_dates(grouped)

        # Prepare PDF in portrait orientation
        pdf = PDF(orientation='P', unit='mm', format='A4')
        pdf.set_header_data(fin_year, header_date, quarter_info)
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # Categorize and prepare data
        ho_units = []
        sof_units = []
        plant_units = []
        
        for idx, ((desc, from_date, to_date), val) in enumerate(grouped.items(), 1):
            auditors_str = ", ".join(val['Auditors'])
            row_data = {
                "Description": desc,
                "Auditors": auditors_str,
                "Manager": val['Manager'] or '',
                "FromDate": from_date.strftime('%d-%m-%Y') if isinstance(from_date, datetime) else str(from_date),
                "ToDate": to_date.strftime('%d-%m-%Y') if isinstance(to_date, datetime) else str(to_date),
                "Remarks": ''
            }
            
            # Categorization logic
            desc_upper = str(desc).strip().upper()
            if desc_upper.startswith("HO-"):
                ho_units.append(row_data)
            elif desc_upper.startswith("SOF-"):
                sof_units.append(row_data)
            else:
                plant_units.append(row_data)

        # Add sections to PDF
        sections = [
            ("HO Units", ho_units),
            ("SOF Units", sof_units),
            ("Plant Units", plant_units)
        ]

        first_section = True
        for section_name, units in sections:
            if units:
                if not first_section:
                    pdf.ln(10)
                
                pdf.add_section_header(section_name)
                pdf.table_header()
                
                for idx, row in enumerate(units, 1):
                    pdf.table_row(idx, row)
                
                first_section = False

        # Generate PDF
        pdf_output = pdf.output(dest='S')
        if isinstance(pdf_output, str):
            pdf_bytes = pdf_output.encode('latin1')
        else:
            pdf_bytes = pdf_output
            
        buffer = io.BytesIO(pdf_bytes)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'audit_programme_{fin_year}.pdf',
            mimetype='application/pdf'
        )

    except Exception as e:
        print(f"Error generating PDF: {e}")
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    app.run(debug=True)
